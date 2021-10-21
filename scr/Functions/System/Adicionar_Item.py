import time
from tkinter import messagebox

from openpyxl import Workbook, load_workbook


def Adicionar_Item(caminho, entrada, janela):
    item = entrada[0:-3]
    coluna = entrada[-2]
    linha = entrada[-1]

    wb = load_workbook(caminho)
    ws = wb.active

    ws[f'{coluna}'+f'{linha}'] = item

    wb.save(caminho)

    time.sleep(1)
    messagebox.showinfo(title='Info', message='Item adicionado com sucesso!')
    janela.destroy()





