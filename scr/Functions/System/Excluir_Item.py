import time
from tkinter import messagebox

from openpyxl import load_workbook


def Excluir_Item(caminho, entrada, janela):
    coluna = entrada[0]
    linha = entrada[1]

    wb = load_workbook(caminho)
    ws = wb.active

    ws[f'{coluna}' + f'{linha}'] = None

    wb.save(caminho)

    time.sleep(1)
    messagebox.showinfo(title='Info', message='Item excluído com sucesso!')
    janela.destroy()
